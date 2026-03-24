import pandas as pd
import openpyxl
import re
import numpy as np
from typing import Dict, List, Tuple, Optional


# Valores válidos en la secuencia de diluciones para CMI según EUCAST: https://www.eucast.org/clinical_breakpoints/
rounded_values = {0.001, 0.002, 0.004, 0.008, 0.016, 0.03, 0.06}

serie_base2 = {2 ** i for i in range(-3, 11)}  # Esto cubre desde 0.125 (2^-3) hasta 1024 (2^10)

VALID_MIC_VALUES = sorted(rounded_values | serie_base2)

def safe_get(row_values, index):
    '''Función auxiliar. Evita IndexError si la fila no tiene todas las columnas, 
    devolviendo None en ese caso.'''
    if index < len(row_values):
        return row_values[index]
    return None

def recortar_a_cmi_valida(numeric_str: str) -> Optional[float]:
    '''Recorta dígitos del final de un string numérico hasta encontrar un valor en VALID_MIC_VALUES.
    Devuelve el valor válido encontrado, o None si no se encuentra ninguno.
    Ejemplo: "0.52" -> 0.5, "0.1252" -> 0.125, "512" -> 512 (ya válido, no se recorta)'''
    s = numeric_str
    while len(s) > 1:
        s = s[:-1]
        try:
            if float(s) in VALID_MIC_VALUES:
                return float(s)
        except ValueError:
            continue
    return None

def parse_notes_column(notes_text: str) -> Dict[str, str]:
    '''
    Parsea el texto de la columna Notes en un diccionario que mapea claves de referencia a sus explicaciones.
    Las claves pueden ser números, letras o combinaciones como "1", "A", "1/A". El texto de las notas se asocia a cada clave
    '''
    if not notes_text or pd.isna(notes_text):
        return {}
    
    notes_dict = {}
    notes_text = str(notes_text)
    
    # Esta expresión captura identificadores de listas compuestos por un número o una letra 
    # (que opcionalmente incluyen una subvariante tras una barra diagonal), seguidos siempre 
    # de un punto y al menos un espacio o tabulación.
    ref_pattern = re.compile(r'(\d+(?:/[A-Z])?|[A-Z](?:/\d+)?)\.[ \t]+')
    refs = []

    for m in ref_pattern.finditer(notes_text):
        ref = m.group(1)
        pos = m.start()
        # Letra sola: válida solo si está en posición de inicio de ítem de lista, es decir:
        # - al inicio del string
        # - precedida por fin de frase: '. ' (punto + espacios)
        # - precedida por doble espacio o más
        # - precedida por salto de línea
        # Esto evita capturar abreviaturas de especie internas como "E." en "E. coli" o
        # "L." en "mg/L. ..." que van precedidas de un solo espacio sin punto previo.
        if re.match(r'^[A-Z]$', ref):
            if pos == 0:
                pass  # inicio del string: siempre válida
            else:
                preceding = notes_text[max(0, pos-4):pos]
                es_inicio_item = (
                    re.search(r'\.\s+$', preceding) or   # fin de frase: ". A."
                    re.search(r'\s{2,}$', preceding) or  # doble espacio o más: "  A."
                    re.search(r'\n\s*$', preceding)       # salto de línea
                )
                if not es_inicio_item:
                    continue  # Es abreviatura de especie u otra mayúscula interna, no referencia
        refs.append((ref, m.start(), m.end()))
    
    # Construir diccionario de notas asociando cada referencia con el 
    # texto que le sigue hasta la siguiente referencia o el final del texto
    for i, (ref, start, end) in enumerate(refs):
        next_start = refs[i+1][1] if i+1 < len(refs) else len(notes_text)
        text = notes_text[end:next_start].strip()
        text = ' '.join(text.split())
        notes_dict[ref] = text
        # Para referencias compuestas como "1/A", también asignamos la misma nota a "1" y "A" por separado
        if '/' in ref:
            for part in ref.split('/'):
                notes_dict[part] = text
    
    return notes_dict


def extract_superscripts_from_value(value) -> List[str]:
    '''Extrae los superíndices de una celda de breakpoints (CMI o halo), que pueden ser letras, números o 
    combinaciones como "1/A", y devuelve una lista de ellos.
    
    Esta función está diseñada para celdas de valores numéricos (columnas de breakpoints),
    NO para el nombre del antibiótico. Para el nombre, usar extract_superscripts_from_antibiotic_name.'''
    if pd.isna(value) or value is None:
        return []
    
    value_str = str(value).strip()
    superscripts = []
    
    note_match = re.search(r'Note([A-Z,\s]+)', value_str, re.IGNORECASE)
    if note_match:
        letters_str = note_match.group(1).replace(',', '').replace(' ', '')
        superscripts.extend(list(letters_str))
        return superscripts
    
    letter_match = re.findall(r'[A-Z]+(?![a-z])', value_str)
    if letter_match:
        for letters in letter_match:
            superscripts.extend(list(letters))
    
    paren_match = re.findall(r'\)(\d+)', value_str)
    if paren_match:
        superscripts.extend(paren_match)
    
    comma_match = re.findall(r',\s*(\d+)', value_str)
    if comma_match:
        superscripts.extend(comma_match)
    
    # Dígitos pegados a letras minúsculas antes de coma o espacio: "Benzylpenicillin2, S."
    inline_match = re.findall(r'[a-z](\d+)(?=[,\s])', value_str)
    if inline_match:
        superscripts.extend(inline_match)
    
    # Superíndices numéricos pegados a un valor CMI: "0.52" -> '2', "0.52,3" -> '2','3'
    # Solo se aplica si el valor es puramente numérico (sin letras mezcladas), porque si hay
    # letras (ej: "14A") ya se capturaron arriba y el recorte generaría falsos positivos
    # ("14A" -> recorta a 1, extrae "4" como superíndice cuando en realidad es parte del número).
    m_mic = re.match(r'^(\d+(?:\.\d+)?)([\d,\s]*)$', value_str)
    if m_mic:
        numeric_str = m_mic.group(1)
        resto = m_mic.group(2)
        try:
            numeric = float(numeric_str)
            if numeric not in VALID_MIC_VALUES:
                valid = recortar_a_cmi_valida(numeric_str)
                if valid is not None:
                    superscripts_str = numeric_str[len(str(valid)):] + resto
                    superscripts_digits = re.findall(r'\d', superscripts_str)
                    superscripts.extend(superscripts_digits)
        except ValueError:
            pass

    name_match = re.search(r'[a-z](\d+)$', value_str, re.IGNORECASE)
    if name_match:
        superscripts.append(name_match.group(1))
    
    return list(dict.fromkeys(superscripts))


def extract_superscripts_from_antibiotic_name(first_cell: str) -> List[str]:
    '''Extrae únicamente los superíndices numéricos del nombre del antibiótico (first_cell).
 
    En las tablas EUCAST los superíndices en el nombre son siempre números (referencias a
    notas numeradas). Las letras nunca son superíndices en el nombre: aparecen solo en las
    celdas de breakpoints (halo de inhibición)
 
    Ejemplos:
        "Meropenem (all indications)1,2"  -> ['1', '2']
        "Ertapenem1"                       -> ['1']
        "Benzylpenicillin2, S. aureus"     -> ['2']'''
    superscripts = []
 
    # Dígitos separados por coma al final: "cillin1,2" o ")1,2"
    comma_match = re.findall(r'(?<=[)a-z\d]),\s*(\d+)', first_cell)
    superscripts.extend(comma_match)
 
    # Dígito(s) al final del string precedidos de espacio, ) o coma
    trailing = re.findall(r'(?<=[\s),])(\d+)(?=\s*(?:,\s*\d+)*\s*$)', first_cell)
    superscripts.extend(trailing)
 
    # Dígito pegado a paréntesis de cierre: ")1", ")1,2"
    paren_digit = re.findall(r'\)(\d+)', first_cell)
    superscripts.extend(paren_digit)
 
    # Dígito pegado directamente al final del nombre: "Ertapenem1", "Benzylpenicillin2"
    name_end_digit = re.search(r'[a-z](\d+)(?:\s*$|(?=,))', first_cell)
    if name_end_digit:
        superscripts.append(name_end_digit.group(1))
 
    return list(dict.fromkeys(superscripts))


def clean_numeric_value(value, is_mic: bool = True) -> Tuple[Optional[float], bool]:
    '''Limpia un valor que se espera numérico (CMI o halo de inhibición) de una tabla EUCAST, 
    detectando también si tiene paréntesis'''
    if pd.isna(value) or value is None:
        return None, False
    
    value_str = str(value).strip()
    
    # valores NO numéricos que indican que no hay un breakpoint definido para esa categoría,
    # o que se debe consultar la nota asociada a ese antibiótico en la tabla.
    null_values = ['-', 'IE', 'NA', 'Note', ''] 
    if value_str in null_values or value_str.startswith('Note'):
        return None, False
    
    has_brackets = False

    if value_str.startswith('(') and ')' in value_str:
        has_brackets = True
        match = re.search(r'\(([^)]+)\)', value_str)
        if match:
            value_str = match.group(1) # sólo nos interesa el valor dentro de los paréntesis para la interpretación y marcar el paréntesis.
    
    cleaned = re.sub(r'[^\d.]', '', value_str)
    
    if not cleaned:
        return None, has_brackets
    
    try:
        numeric_value = float(cleaned) # valor en formato numérico

        # si es CMI y el valor numérico no es válido, intentar recortar dígitos del final para encontrar un valor válido en la secuencia de diluciones CMIs
        if is_mic and numeric_value and numeric_value not in VALID_MIC_VALUES: 
            valid = recortar_a_cmi_valida(str(int(numeric_value)) if numeric_value >= 10 else cleaned)
            if valid is not None:
                numeric_value = valid
        
        return numeric_value, has_brackets
    
    except ValueError:
        return None, has_brackets


def parse_atu_value(value) -> Tuple[Optional[float], Optional[float]]:
    '''Parsea un valor de ATU (Área de Transición de Incertidumbre) que puede estar en formato "min-max" o un solo valor
    Devuelve una tupla (min, max) con los valores numéricos, o (None, None) si no se encuentra un valor válido.'''
    if pd.isna(value) or value is None:
        return None, None
    
    value_str = str(value).strip()
    
    if value_str in ['-', 'IE', 'NA', '', 'Note'] or value_str.startswith('Note'):
        return None, None
    
    # En algunos casos el ATU se expresa como un rango "min-max"
    match = re.search(r'(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)', value_str)
    if match:
        try:
            min_val = float(match.group(1))
            max_val = float(match.group(2))
            return min_val, max_val
        except ValueError:
            return None, None
    
    try:
        # Para casos de valor único, se intenta limpiar y convertir a número. Si no es un número válido, se devuelve None.
        cleaned = re.sub(r'[^\d.]', '', value_str)
        if cleaned:
            single_val = float(cleaned)
            return single_val, single_val
    except ValueError:
        return None, None
    
    return None, None


def extract_breakpoints_from_row(row_values: List) -> Dict:
    '''Extrae los breakpoints de una fila de la tabla, limpiando los valores y detectando si hay paréntesis.
    Devuelve un diccionario con los breakpoints limpios y un indicador de si hay valores entre paréntesis.'''
    breakpoints = {}
    
    has_both = len(row_values) >= 8 # Tablas completas con CMI y halos de inhibición. 
                                    # Para ciertos microorganismos, como N. gonorrhoeae esto no sucede
    
    if has_both:
        mic_s, mic_s_brackets = clean_numeric_value(safe_get(row_values, 1), is_mic=True)
        mic_r, mic_r_brackets = clean_numeric_value(safe_get(row_values, 2), is_mic=True)
        atu_mic_min, atu_mic_max = parse_atu_value(safe_get(row_values, 3))
        
        zone_s, zone_s_brackets = clean_numeric_value(safe_get(row_values, 5), is_mic=False)
        zone_r, zone_r_brackets = clean_numeric_value(safe_get(row_values, 6), is_mic=False)
        atu_zone_min, atu_zone_max = parse_atu_value(safe_get(row_values, 7))
    else: 
        # tablas con sólo CMI, como las de N. gonorrhoeae
        val_s, val_s_brackets = clean_numeric_value(safe_get(row_values, 1), is_mic=True)
        val_r, val_r_brackets = clean_numeric_value(safe_get(row_values, 2), is_mic=True)
        atu_min, atu_max = parse_atu_value(safe_get(row_values, 3))
        
        mic_s, mic_s_brackets = val_s, val_s_brackets
        mic_r, mic_r_brackets = val_r, val_r_brackets
        atu_mic_min, atu_mic_max = atu_min, atu_max
        
        zone_s, zone_s_brackets = None, False
        zone_r, zone_r_brackets = None, False
        atu_zone_min, atu_zone_max = None, None
    
    breakpoints['MIC_S'] = mic_s
    breakpoints['MIC_R'] = mic_r
    breakpoints['ATU_MIC_min'] = atu_mic_min
    breakpoints['ATU_MIC_max'] = atu_mic_max
    breakpoints['Zone_S'] = zone_s
    breakpoints['Zone_R'] = zone_r
    breakpoints['ATU_Zone_min'] = atu_zone_min
    breakpoints['ATU_Zone_max'] = atu_zone_max
    
    has_brackets = mic_s_brackets or mic_r_brackets or zone_s_brackets or zone_r_brackets
    breakpoints['brackets'] = 1 if has_brackets else 0 # indicador binario de si hay valores entre paréntesis en alguno de los breakpoints, lo cual implica que se deben consultar las notas asociadas a ese antibiótico en la tabla.
    
    return breakpoints


def parse_antibiotic_name(name: str) -> Tuple[str, Optional[str], Optional[str], Optional[str]]:
    '''Parsea la fila donde está el nombre del antibiótico para extraer: nombre base (limpio, siempre presente), 
    vía de administración, indicación y aplicación a especies (estas últimas 3 pueden ser None si no se encuentran)'''

    if not isinstance(name, str) or not name.strip():
        return None, None, None, None
    
    name = name.replace('\xa0', ' ') # Limpiar espacios no separables para evitar problemas al parsear.
    name = ' '.join(name.split())
    
    nombre_base = name
    via_administracion = None
    indicacion = None
    aplicacion_especies = None

    # Limpiar superíndices numéricos separados por coma ANTES de cualquier split
    # Ej: "Meropenem (all indications)1,2" -> "Meropenem (all indications)"
    # Solo elimina dígito,dígito al final de la cadena, no antes de especies
    nombre_base = re.sub(r'(\d+)(,\d+)+(?=\s*$)', '', nombre_base).strip()
    # También limpiar superíndice suelto pegado a paréntesis de cierre: ")1" -> ")"
    nombre_base = re.sub(r'(\))([\d¹²³⁴⁵⁶⁷⁸⁹⁰]+)', r'\1', nombre_base).strip()

    # Palabras clave para detectar aplicación de especies
    palabras_clave_especie = [
        'except', 'enterobacterales', 'morganellaceae', 'excluding', 'including',
        'other than', 'streptococc', 'e. coli', 'klebsiella', 'proteus', 'salmonella',
        'pseudomonas', 'acinetobacter', 'citrobacter', 'raoultella', 'enterobacter',
        'serratia', 'morganella', 'staphylococ', 'lugdunensis', 'haemophilus',
        'moraxella', 'stenotrophomonas', 'bacteroides', 'clostridium', 'fusobacterium',
        'prevotella', 'peptostreptococcus', 'neisseria', 'listeria', 'corynebacterium',
        'nocardia', 'burkholderia', 'chryseobacterium', 'elizabethkingia',
        'other enterococci', 'other streptococci', 'other staphylococci',
        'groups', 'group ',                                         
    ]

    def _es_aplicacion_especie(texto: str) -> bool:
        t = texto.lower()
        if any(kw in t for kw in palabras_clave_especie):
            return True
        
        # Abreviatura: S. aureus o S.aureus (insensible a espacios)
        if re.match(r'^[A-Z]\.\s*[a-z]+', texto):
            return True
        
        # Género completo en mayúscula seguido de nombre en minúsculas: "Streptococcus spp"
        if re.match(r'^[A-Z][a-z]+ [a-z]+', texto):
            return True

        # Acrónimos específicos
        acronimos_comunes = {'mrsa', 'mssa', 'vrsa', 'vre', 'kpc', 'visa', 'esbl', 'mdr'}
        if t in acronimos_comunes:
            return True
            
        return False

    # La aplicación de especies va siempre detrás de la primera coma en la celda con 
    # el nombre del antibiótico, incluso si hay paréntesis con indicaciones.
    # IMPORTANTE: solo se considera la coma si está FUERA de paréntesis, para evitar
    # que comas dentro de indicaciones como "(all indications, including meningitis)"
    # se interpreten como separador de aplicacion_especies.
    def _primera_coma_fuera_parentesis(text):
        depth = 0
        for i, ch in enumerate(text):
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif ch == ',' and depth == 0:
                return text[:i].strip(), text[i+1:].strip()
        return text.strip(), None

    parte_antes, parte_despues = _primera_coma_fuera_parentesis(nombre_base)
    if parte_despues is not None and _es_aplicacion_especie(parte_despues):
        nombre_base = parte_antes
        aplicacion_especies = parte_despues
        aplicacion_especies = re.sub(r'[\d¹²³⁴⁵⁶⁷⁸⁹⁰]+$', '', aplicacion_especies).strip()

    # Separar género bacteriano sin coma: "Cloxacillin Streptococcus groups A, C and G"
    # Solo actúa si no hay ya aplicacion_especies y el candidato empieza por un género
    # conocido (palabra de >=5 letras en mayúscula inicial) o acrónimo clínico
    if aplicacion_especies is None:
        match_genus = re.search(r'\s+([A-Z][a-z]{4,}\b.*)', nombre_base)
        if match_genus:
            candidato = match_genus.group(1).strip()
            if _es_aplicacion_especie(candidato):
                nombre_base = nombre_base[:match_genus.start()].strip() # nombre del antibiótico
                aplicacion_especies = candidato

    # Paréntesis (incluso sin cerrar): indicaciones. Hay algunos paréntesis 
    # en las celdas de las tablas de EUCAST que no se llegaron a cerrar
    indicaciones = []
    # Paréntesis cerrados
    for m in re.finditer(r'\(([^)]+)\)', nombre_base):
        indicaciones.append(m.group(1).strip())
    # Paréntesis sin cerrar al final -> solo si NO hay ya paréntesis cerrados
    if not indicaciones:
        unclosed = re.search(r'\(([^)]+)$', nombre_base)
        if unclosed:
            indicaciones.append(unclosed.group(1).strip())

    # Filtrar indicaciones que significan "todas" (restricciones sólo para ciertas indicaciones)
    _todas = {'all indications', 'all indication', 'all', 'any indication'}
    indicaciones = [i for i in indicaciones if i.lower().strip() not in _todas]

    if indicaciones:
        indicacion = indicaciones[-1] if len(indicaciones) > 1 else indicaciones[0]

    # Eliminar paréntesis cerrados del nombre base
    nombre_base = re.sub(r'\s*\([^)]+\)', '', nombre_base).strip()
    # Eliminar paréntesis sin cerrar al final
    nombre_base = re.sub(r'\s*\([^)]*$', '', nombre_base).strip()
    nombre_base = ' '.join(nombre_base.split())

    # Limpiar letras superíndice sueltas al final: "Piperacillin-tazobactam A" -> "Piperacillin-tazobactam"
    # o múltiples letras separadas por coma: "Cefotaxime B,C" -> "Cefotaxime"
    nombre_base = re.sub(r'\s+[A-Z](?:\s*,\s*[A-Z])*\s*$', '', nombre_base).strip()

    # Limpiar comas y superíndices numéricos al final
    nombre_base = nombre_base.rstrip(', \t')
    nombre_base = re.sub(r'[\d¹²³⁴⁵⁶⁷⁸⁹⁰]+$', '', nombre_base).strip()
    nombre_base = nombre_base.rstrip(', \t')

    # Vía de administración, al final del nombre con posibilidad iv u oral, incluso con espacios: 
    # "Amoxicillin iv", "Amoxicillin oral", etc.
    vias_patron = r'\s+(iv|oral)$'
    match_via = re.search(vias_patron, nombre_base, re.IGNORECASE)
    if match_via:
        via_administracion = match_via.group(1).lower()
        nombre_base = re.sub(vias_patron, '', nombre_base, flags=re.IGNORECASE).strip()

    return nombre_base, via_administracion, indicacion, aplicacion_especies


def extract_antibiotics_from_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Extrae antibióticos de una hoja seleccionada de un archivo EUCAST.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: String con el nombre de la hoja a procesar
    Returns:
        DataFrame con columnas: grupo, antibiotico, via_administracion, indicacion, aplicacion_especies,
        MIC_S, MIC_R, ATU_MIC_min, ATU_MIC_max, Zone_S, Zone_R, ATU_Zone_min, ATU_Zone_max, notes
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Celdas combinadas de las tablas en la columna A (grupos de antibióticos)
    merged_ranges = list(ws.merged_cells.ranges)
    
    # Celdas combinadas de la columna A para identificar filas de header y grupos
    merged_cells_col_a = {}
    for mr in merged_ranges:
        if mr.min_col == 1 and mr.max_col == 1: # solo nos interesan los merges de la columna A
            for row in range(mr.min_row, mr.max_row + 1):
                merged_cells_col_a[row] = (mr.min_row, mr.max_row)
    
    # Filas de los header de las tablas donde encontramos los nombres de los grupos EUCAST (ej: "Penicillins", "Carbapenems", etc.)
    header_rows = sorted(set(start for start, end in merged_cells_col_a.values()))

    # Detectar columna Notes
    notes_col_index = None
    for header_row in header_rows:
        # buscamos la columna que contiene "Notes" en la fila del header, ignorando mayúsculas y espacios
        for col in range(1, ws.max_column + 1):
            value = ws.cell(header_row, col).value
            # como Notes puedde tener más texto, incluso superíndice, buscamos que empiece por "notes"
            if isinstance(value, str) and value.strip().lower().startswith("notes"):
                notes_col_index = col 
                break
        if notes_col_index:
            break

    if notes_col_index is None:
        raise ValueError("No se encontró la columna Notes")

    # Construir merged_notes_top solo con combinaciones de celdas verticales en Notes.
    # Normalmente cada grupo tiene un bloque de notas combinadas debajo del header, 
    # ocupando todo el espacio de la columna
    merged_notes_top = {}
    for mr in merged_ranges:
        if mr.min_col == notes_col_index and mr.max_col == notes_col_index:
            for row in range(mr.min_row, mr.max_row + 1):
                merged_notes_top[row] = mr.min_row


    # Función auxiliar para obtener notas de cualquier fila
    def get_notes_for_row(header_row_num):
        for row in range(header_row_num + 1, header_row_num + 100):
            if row in merged_notes_top and merged_notes_top[row] == row:
                # esta fila es el inicio de un bloque de notas combinadas, obtenemos el texto de esa celda
                notes_text = ws.cell(row=row, column=notes_col_index).value
                if notes_text and not str(notes_text).strip().lower().startswith("notes"):
                    return parse_notes_column(notes_text) # diccionario de notas para esa tabla
        return {}

    # Leer la hoja completa con pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    antibiotics_data = []
    current_group = None
    in_table = False
    empty_rows_count = 0
    current_notes_dict = {}

    # Recorrer fila por fila para detectar grupos, antibióticos y asociar notas. 
    # No se puede hacer todo con pandas porque la estructura de las tablas es compleja (grupos con filas combinadas, 
    # notas en bloques combinados debajo del header, etc.)
    for idx, row in df.iterrows():
        row_num = idx + 1
        row_values = row.tolist()
        first_cell = row_values[0] if pd.notna(row_values[0]) else ""
        first_cell = str(first_cell).strip()

        is_empty_row = all(pd.isna(val) or str(val).strip() == "" for val in row_values)

        if is_empty_row:
            empty_rows_count += 1
            # Si encontramos 2 filas vacías seguidas dentro de una tabla, asumimos que la tabla ha terminado. 
            # A menudo hay dos filas vacías entre tablas
            if empty_rows_count >= 2 and in_table:
                in_table = False # reseteamos estado para buscar la siguiente tabla y grupo de antibióticos
                current_group = None
                current_notes_dict = {}
            continue
        else:
            empty_rows_count = 0

        # Detectar nuevo grupo de antibióticos por fila combinada en la columna A. Estas filas combinadas suelen 
        # contener el nombre del grupo EUCAST (ej: "Penicillins", "Carbapenems", etc.)
        if row_num in merged_cells_col_a:
            merge_start, merge_end = merged_cells_col_a[row_num]
            if merge_start == row_num and first_cell:
                
                # Algunas filas combinadas no son grupos de antibióticos, sino que contienen texto explicativo o notas.
                # Para evitar confundirlas con grupos, podemos excluir aquellas que contengan ciertas palabras clave
                palabras_excluir = ['expert rules', 'mic determination', 'enterobacterales*',
                                     'medium:', 'inoculum:', 'incubation:', 'reading:',
                                     'quality control:', 'recent taxonomic', 'neisseria gonorrhoeae',
                                     'for comments on dosages', 'disk diffusion criteria']
                
                if not any(palabra in first_cell.lower() for palabra in palabras_excluir):
                    # Limpiar superíndices numéricos al final del nombre del grupo: "Penicillins1" -> "Penicillins"
                    # Dejo los caracteres de superíndices aunque la transcripción del texto en celda los obtiene 
                    # como dígitos normales.
                    current_group = re.sub(r'[\d¹²³⁴⁵⁶⁷⁸⁹⁰]+$', '', first_cell).strip()
                    in_table = True

                    # Obtener notas asociadas al grupo
                    current_notes_dict = get_notes_for_row(row_num + 1)

                    continue

        # Procesar filas de antibióticos
        if in_table and current_group:

            if row_num not in merged_cells_col_a and first_cell:
                # Pasar del encabezado de la tabla a las filas de antibióticos.
                if any(x in first_cell for x in ['S ≤', 'R >', '≤', '≥', 'ATU']):
                    continue
                
                # Nombre del antibiótico, vía de administración, indicación y aplicación a especies 
                nombre_base, via_administracion, indicacion, aplicacion_especies = parse_antibiotic_name(first_cell)

                if nombre_base and nombre_base != '-':
                    # Extraer breakpoints de la fila, limpiando los valores y detectando si hay paréntesis.
                    breakpoints = extract_breakpoints_from_row(row_values)

                    superscripts = set()

                    # Extraer superíndices de las celdas de breakpoints. Para tablas completas con CMI y halos, 
                    # se extraen de las columnas 1, 2, 5 y 6. Para tablas con sólo CMI, sólo de las columnas 1 y 2.
                    if len(row_values) >= 8:
                        for col_idx in [1, 2, 5, 6]:
                            val = safe_get(row_values, col_idx)
                            superscripts.update(extract_superscripts_from_value(val))
                    else:
                        for col_idx in [1, 2]:
                            val = safe_get(row_values, col_idx)
                            superscripts.update(extract_superscripts_from_value(val))

                    # Extraer superíndices del nombre del antibiótico usando la función especializada.
                    # Se usa extract_superscripts_from_antibiotic_name (NO extract_superscripts_from_value)
                    # para evitar capturar siglas internas del nombre como CNS, ATU, etc. como si fueran
                    # referencias a notas. Solo captura letras/números aislados al final del nombre o
                    # pegados a paréntesis de cierre.
                    superscripts.update(extract_superscripts_from_antibiotic_name(first_cell))

                    note_texts = []
                    for sup in sorted(superscripts):
                        if sup in current_notes_dict:
                            note_texts.append(current_notes_dict[sup])

                    unique_notes = list(dict.fromkeys(note_texts))

                    # Combinar las notas únicas en un solo string separado por salto de línea, o None si no hay notas.
                    notes_combined = ' \n'.join(unique_notes) if unique_notes else None

                    # Añadir el antibiótico y sus datos como diccionario a la lista, 
                    # incluyendo el grupo actual, los breakpoints y las notas combinadas.
                    antibiotics_data.append({
                        'grupo': current_group,
                        'antibiotico': nombre_base,
                        'via_administracion': via_administracion,
                        'indicacion': indicacion,
                        'aplicacion_especies': aplicacion_especies,
                        **breakpoints,
                        'notes': notes_combined
                    })

    result_df = pd.DataFrame(antibiotics_data) # DataFrame final con los antibióticos extraídos de la hoja, con sus datos asociados, incluyendo el grupo EUCAST, vía de administración, indicación, aplicación a especies, breakpoints y notas.
    return result_df


def extract_all_antibiotics(file_path: str, sheet_selection=None, version: str = None) -> pd.DataFrame:
    """
    Extrae antibióticos de las hojas seleccionadas del archivo EUCAST.

    Args:
        file_path: Ruta al archivo Excel
        sheet_selection: Puede ser:
            - None: Procesa TODAS las hojas
            - Lista de nombres: ['Enterobacterales', 'N.gonorrhoeae']
            - Lista de índices: [0, 1, 2]
            - Rango (tuple): (0, 3) procesa hojas del índice 0 al 2 (exclusivo)
            - String 'all': Procesa todas las hojas (equivalente a None)
        version: String con la versión de las tablas EUCAST (ej: 'v14.0', '2024')

    Returns:
        DataFrame con columnas: version, grupo_EUCAST, grupo, antibiotico, via_administracion, indicacion, aplicacion_especies,
        brackets, MIC_S, MIC_R, ATU_MIC_min, ATU_MIC_max, Zone_S, Zone_R, ATU_Zone_min, ATU_Zone_max, notes
    """
    xl_file = pd.ExcelFile(file_path)

    # nombres de las hojas del archivo Excel
    all_sheet_names = xl_file.sheet_names
    
    sheets_to_process = []
    
    # si se utiliza manualmente, puede ser de ayuda tener esta posibilidad
    if sheet_selection is None or sheet_selection == 'all':
        sheets_to_process = all_sheet_names

    # también para modos manuales, puede ser útil pasar una lista de nombres o índices
    elif isinstance(sheet_selection, list):
        if all(isinstance(s, str) for s in sheet_selection): # nombres
            sheets_to_process = [s for s in sheet_selection if s in all_sheet_names]
            not_found = [s for s in sheet_selection if s not in all_sheet_names]
            if not_found:
                print(f"Advertencia: Hojas no encontradas: {not_found}")
        
        elif all(isinstance(s, int) for s in sheet_selection): # índices
            sheets_to_process = [all_sheet_names[i] for i in sheet_selection 
                               if 0 <= i < len(all_sheet_names)]
            invalid = [i for i in sheet_selection if i < 0 or i >= len(all_sheet_names)]
            if invalid:
                print(f"Advertencia: Índices fuera de rango: {invalid} (total hojas: {len(all_sheet_names)})")
        
        else:
            raise ValueError("La lista debe contener solo strings (nombres) o solo integers (índices)")
    
    # habitualmente el endpoint mandará la tupla con página inicial y final
    elif isinstance(sheet_selection, tuple) and len(sheet_selection) == 2:
        start, end = sheet_selection
        sheets_to_process = all_sheet_names[start:end]
    
    else:
        raise ValueError("sheet_selection debe ser None, 'all', una lista, o una tupla (start, end)")
    
    if not sheets_to_process:
        raise ValueError("No hay hojas para procesar. Verifica la selección.")
    
    # inicialización del contenedor de todos los datos extraídos de las hojas seleccionadas
    all_data = []
    
    print(f"Hojas disponibles en el archivo ({len(all_sheet_names)}):")
    
    for i, name in enumerate(all_sheet_names):
        marker = "✓" if name in sheets_to_process else " "
        print(f"  [{marker}] {i}: {name}")
    
    if version:
        print(f"\nVersión: {version}")
    
    print(f"\nProcesando {len(sheets_to_process)} hoja(s)...")
    print("="*80)
    
    # procesa cada hoja seleccionada, extrayendo los antibióticos y sus datos, y añadiendo columnas de grupo y versión
    for sheet_name in sheets_to_process:
        print(f"\nProcesando hoja: {sheet_name}")

        try:
            df_sheet = extract_antibiotics_from_sheet(file_path, sheet_name)
            df_sheet['grupo_EUCAST'] = sheet_name # añadimos columna con el grupo EUCAST, que se corresponde con el nombre de la hoja
            df_sheet['version'] = version # la versión que venga en el argumento, vendrá por formulario.
            all_data.append(df_sheet)
            print(f"  ✓ Encontrados {len(df_sheet)} antibióticos")

        except Exception as e:
            print(f"  ✗ Error procesando {sheet_name}: {str(e)}")
            continue
    
    if not all_data:
        raise ValueError("No se pudo extraer datos de ninguna hoja")
    
    result = pd.concat(all_data, ignore_index=True) # encadenamos los datos de todas las hojas seleccionadas en un solo DataFrame
    
    column_order = [
        'version', 'grupo_EUCAST', 'grupo', 'antibiotico', 'via_administracion', 'indicacion', 'aplicacion_especies',
        'brackets',
        'MIC_S', 'MIC_R', 'ATU_MIC_min', 'ATU_MIC_max',
        'Zone_S', 'Zone_R', 'ATU_Zone_min', 'ATU_Zone_max',
        'notes'
    ]
    result = result[column_order] # reordenamos columnas para que tengan un orden lógico y consistente
    
    # enteros para estas columnas de halos de inhibición y ATU de zona
    int_columns = ['Zone_S', 'Zone_R', 'ATU_Zone_min', 'ATU_Zone_max']
    for col in int_columns:
        if col in result.columns:
            result[col] = result[col].astype('Int64')
    
    return result